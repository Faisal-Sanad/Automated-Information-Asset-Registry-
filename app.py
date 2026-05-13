from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import db
import psycopg2
import os
import csv
import io
import re
import json
import bcrypt
import jwt as pyjwt
from datetime import datetime, timedelta, timezone
from functools import wraps
from audit_log import log_change
import threading
from dotenv import load_dotenv
from email_service import (send_2fa_code, send_registration_approval,
                            send_registration_rejection,
                            send_registration_received,
                            send_admin_registration_notification,
                            send_pending_change_notification,
                            send_password_reset_code,
                            send_admin_password_reset_notification,
                            send_account_deleted_notification,
                            generate_otp)

# Load environment variables explicitly rather than relying on the
# side-effect of importing db.py. This makes JWT_SECRET resolution
# deterministic regardless of import order and ensures the app fails
# at startup if the .env file is missing.
load_dotenv()

app = Flask(__name__)
app.config['SESSION_COOKIE_SECURE'] = True

# JWT secret must be configured in the environment. A hard-coded fallback
# would be readable in the source repository, allowing anyone to forge
# valid tokens and impersonate administrators. Fail fast at startup so a
# misconfigured deployment is obvious rather than silently insecure.
JWT_SECRET = os.getenv("JWT_SECRET")
if not JWT_SECRET:
    raise RuntimeError(
        "JWT_SECRET environment variable is not set. "
        "Generate one with: python -c \"import secrets; print(secrets.token_urlsafe(48))\" "
        "and add it to your .env file as JWT_SECRET=<value>."
    )
if len(JWT_SECRET) < 32:
    raise RuntimeError(
        "JWT_SECRET is too short to be cryptographically secure. "
        "Use at least 32 characters (ideally 48+). "
        "Generate a new one with: python -c \"import secrets; print(secrets.token_urlsafe(48))\"."
    )
JWT_EXPIRY_HOURS = 8

CORS(app, supports_credentials=True, origins=['http://localhost:5000', 'http://127.0.0.1:5000'])


# ── Allow-list of asset columns that may be dynamically updated ───────────
# Any code path that interpolates a column name into a SQL statement MUST
# validate against this set first. Prevents SQL injection via user-supplied
# field names (e.g. a tampered reconciliation detail string or a crafted
# pending_changes.field_changed entry).
UPDATABLE_ASSET_COLUMNS = frozenset({
    'system_name', 'data_category', 'description', 'business_function',
    'data_owner', 'custodian', 'data_classification', 'data_location',
    'who_has_access', 'access_mechanism', 'security_controls',
    'encryption_in_transit', 'encryption_at_rest',
    'retention_period', 'backup_retention', 'disposal_method',
    'primary_use', 'policy_reference', 'control_mapping',
    'last_review_date', 'remarks',
    'data_sensitivity', 'business_criticality', 'risk_rating', 'key_risks',
    # Risk-scoring fields added per industry mentor feedback (Apr 2026)
    'inherent_likelihood', 'inherent_impact',
    'residual_likelihood', 'residual_impact',
    'risk_treatment_plan', 'review_frequency', 'target_review_date',
    'project_manager', 'network_ports_protocols'
})


# ── Form value normalisation ──────────────────────────────────────────────
def clean(value):
    """
    Coerce a submitted form value into a database-friendly form.

    HTML forms submit unset fields as empty strings rather than nulls,
    but several asset columns carry CHECK constraints (e.g.
    encryption_at_rest must be 'Yes'/'No'/'Partial') that reject empty
    strings. Without coercion, an INSERT fails with a CheckViolation.
    Returns None when the value is None, empty, or whitespace-only.
    Strips surrounding whitespace from non-empty strings.
    """
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped if stripped else None


def clean_int(value, min_val=None, max_val=None):
    """
    Coerce a submitted value to an integer or None.

    Returns None if the value is empty, whitespace, unparseable, or
    falls outside the optional [min_val, max_val] range. Used for
    risk-scoring fields where the database CHECK constraint requires
    1-5 but the frontend may submit empty strings or out-of-range values.
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == '' or s.lower() == 'none':
        return None
    try:
        n = int(s)
    except (ValueError, TypeError):
        return None
    if min_val is not None and n < min_val:
        return None
    if max_val is not None and n > max_val:
        return None
    return n


# ── Mandatory fields for CBB OM-5.5 / ISO 27001 A.5.9 compliance ──────────
# Canonical definition used by every compliance score calculation in the
# system. Must match MANDATORY_FIELDS in reconciliation.py and the
# compliance_dashboard view in Supabase. An asset is Compliant only when
# all MANDATORY_FIELD_COUNT fields are populated (strict 100% threshold).
MANDATORY_ASSET_FIELDS = [
    'asset_id', 'system_name', 'data_category', 'data_owner',
    'data_classification', 'encryption_in_transit', 'encryption_at_rest',
    'retention_period', 'control_mapping'
]
MANDATORY_FIELD_COUNT = len(MANDATORY_ASSET_FIELDS)  # 9

# SQL fragment evaluating how many of the mandatory fields are populated
# for each asset. Treats NULL, empty string and whitespace-only as missing.
MANDATORY_SCORE_SQL = "(" + " + ".join(
    f"CASE WHEN a.{f} IS NOT NULL AND TRIM(COALESCE(a.{f}::text, '')) != '' THEN 1 ELSE 0 END"
    for f in MANDATORY_ASSET_FIELDS
) + ")"


# ── Background email helper ───────────────────────────────────────────────
def send_email_async(fn, *args, **kwargs):
    """Run an email function in a background thread so it never blocks."""
    t = threading.Thread(target=fn, args=args, kwargs=kwargs, daemon=True)
    t.start()


# ── JWT helpers ────────────────────────────────────────────────────────────
def generate_token(username, role):
    payload = {
        "sub": username,
        "role": role,
        "iat": datetime.now(timezone.utc),
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS)
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm="HS256")


def decode_token(token):
    return pyjwt.decode(token, JWT_SECRET, algorithms=["HS256"])


def get_token_from_request():
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:]
    return None


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = get_token_from_request()
        if not token:
            return jsonify({'error': 'Unauthorised — no token'}), 401
        try:
            payload = decode_token(token)
        except pyjwt.ExpiredSignatureError:
            return jsonify({'error': 'Token expired'}), 401
        except pyjwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401
        request.current_user = payload["sub"]
        request.current_role = payload["role"]
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = get_token_from_request()
        if not token:
            return jsonify({'error': 'Unauthorised — no token'}), 401
        try:
            payload = decode_token(token)
        except pyjwt.ExpiredSignatureError:
            return jsonify({'error': 'Token expired'}), 401
        except pyjwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401
        if payload.get("role") != "admin":
            return jsonify({'error': 'Forbidden — admin only'}), 403
        request.current_user = payload["sub"]
        request.current_role = payload["role"]
        return f(*args, **kwargs)
    return decorated


# ── Auth endpoints ─────────────────────────────────────────────────────────
@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip().encode('utf-8')

    cols, rows = db.fetch_all(
        "SELECT username, password_hash, role, full_name, email, tfa_enabled, is_approved FROM users WHERE username = %s AND is_active = TRUE",
        (username,)
    )
    if not rows:
        return jsonify({'error': 'Invalid username or password'}), 401

    _, password_hash, role, full_name, email, tfa_enabled, is_approved = rows[0]

    if not bcrypt.checkpw(password, password_hash.encode('utf-8')):
        return jsonify({'error': 'Invalid username or password'}), 401

    if not is_approved:
        return jsonify({'error': 'Account pending administrator approval'}), 403

    # If 2FA enabled, send OTP and return partial auth response
    if tfa_enabled and email:
        otp = generate_otp()
        expires = datetime.now(timezone.utc) + timedelta(minutes=5)
        db.execute(
            "UPDATE users SET tfa_code = %s, tfa_expires_at = %s WHERE username = %s",
            (otp, expires, username)
        )
        send_email_async(send_2fa_code, email, otp, username)
        return jsonify({'requires_2fa': True, 'username': username})

    # Update last_login timestamp
    db.execute("UPDATE users SET last_login = NOW() WHERE username = %s", (username,))
    log_change(asset_id='SYSTEM', action='LOGIN', changed_by=username,
               notes=f'Login from {request.remote_addr}')

    token = generate_token(username, role)
    return jsonify({'token': token, 'role': role, 'name': full_name})


@app.route('/api/logout', methods=['POST'])
def logout():
    # JWT is stateless — client discards the token
    return jsonify({'success': True})


# ── 2FA Verification ───────────────────────────────────────────────────────
@app.route('/api/login/verify-2fa', methods=['POST'])
def verify_2fa():
    data = request.json
    username = data.get('username', '').strip()
    code     = data.get('code', '').strip()

    cols, rows = db.fetch_all(
        "SELECT username, role, full_name, tfa_code, tfa_expires_at FROM users WHERE username = %s AND is_active = TRUE",
        (username,)
    )
    if not rows:
        return jsonify({'error': 'Invalid request'}), 401

    _, role, full_name, stored_code, expires_at = rows[0]

    if not stored_code or code != stored_code:
        return jsonify({'error': 'Invalid verification code'}), 401

    if expires_at and datetime.now(timezone.utc) > expires_at.replace(tzinfo=timezone.utc):
        return jsonify({'error': 'Verification code has expired'}), 401

    # Clear the OTP and update last login
    db.execute(
        "UPDATE users SET tfa_code = NULL, tfa_expires_at = NULL, last_login = NOW() WHERE username = %s",
        (username,)
    )
    log_change(asset_id='SYSTEM', action='LOGIN', changed_by=username,
               notes=f'2FA login from {request.remote_addr}')

    token = generate_token(username, role)
    return jsonify({'token': token, 'role': role, 'name': full_name})


# ── Registration ───────────────────────────────────────────────────────────
@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    username  = data.get('username', '').strip()
    password  = data.get('password', '').strip()
    full_name = data.get('full_name', '').strip()
    email     = data.get('email', '').strip()

    if not username or not password or not email:
        return jsonify({'error': 'Username, password, and email are required'}), 400

    if not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        return jsonify({'error': 'Invalid email address'}), 400

    # Check username not already taken
    _, rows = db.fetch_all("SELECT username FROM users WHERE username = %s", (username,))
    if rows:
        return jsonify({'error': 'Username already exists'}), 409

    password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    db.execute("""
        INSERT INTO users (username, password_hash, role, full_name, email, is_approved)
        VALUES (%s, %s, 'viewer', %s, %s, FALSE)
    """, (username, password_hash, full_name, email))

    # Email 1: Confirm receipt to the registering user
    send_email_async(send_registration_received, email, username)

    # Email 2: Notify all admins of new registration request
    _, admin_rows = db.fetch_all(
        "SELECT email, full_name FROM users WHERE role = 'admin' AND is_active = TRUE AND email IS NOT NULL"
    )
    for row in admin_rows:
        if row[0]:
            send_email_async(send_admin_registration_notification, row[0], username, full_name, email)

    return jsonify({'success': True, 'message': 'Registration submitted. Awaiting administrator approval.'})


# ── Pending registrations (admin) ──────────────────────────────────────────
@app.route('/api/registrations', methods=['GET'])
@admin_required
def list_registrations():
    cols, rows = db.fetch_all(
        "SELECT user_id, username, full_name, email, created_at FROM users WHERE is_approved = FALSE AND is_active = TRUE ORDER BY created_at"
    )
    return jsonify([dict(zip(cols, [str(v) if v is not None else '' for v in row])) for row in rows])


@app.route('/api/registrations/<username>/approve', methods=['POST'])
@admin_required
def approve_registration(username):
    db.execute("UPDATE users SET is_approved = TRUE WHERE username = %s", (username,))
    _, rows = db.fetch_all("SELECT email, full_name FROM users WHERE username = %s", (username,))
    if rows and rows[0][0]:
        send_email_async(send_registration_approval, rows[0][0], username)
    log_change(asset_id='SYSTEM', action='UPDATE', field_changed='is_approved',
               old_value='False', new_value='True',
               changed_by=request.current_user,
               notes=f"User '{username}' registration approved")
    return jsonify({'success': True})


@app.route('/api/registrations/<username>/reject', methods=['POST'])
@admin_required
def reject_registration(username):
    reason = (request.json or {}).get('reason', '')
    db.execute("UPDATE users SET is_active = FALSE WHERE username = %s AND is_approved = FALSE", (username,))
    _, rows = db.fetch_all("SELECT email, full_name FROM users WHERE username = %s", (username,))
    if rows and rows[0][0]:
        send_email_async(send_registration_rejection, rows[0][0], username, reason)
    log_change(asset_id='SYSTEM', action='DELETE', field_changed='registration',
               changed_by=request.current_user,
               notes=f"User '{username}' registration rejected. Reason: {reason}")
    return jsonify({'success': True})


@app.route('/api/me')
def me():
    token = get_token_from_request()
    if not token:
        return jsonify({'logged_in': False})
    try:
        payload = decode_token(token)
        return jsonify({'logged_in': True, 'user': payload['sub'], 'role': payload['role']})
    except pyjwt.InvalidTokenError:
        return jsonify({'logged_in': False})


# ── Dashboard stats ────────────────────────────────────────────────────────
@app.route('/api/stats')
def stats():
    _, rows = db.fetch_all("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN compliance_status = 'Compliant' THEN 1 ELSE 0 END) as compliant,
               SUM(CASE WHEN compliance_status = 'Non-Compliant' THEN 1 ELSE 0 END) as non_compliant,
               SUM(CASE WHEN review_status = 'Review Overdue' THEN 1 ELSE 0 END) as overdue
        FROM compliance_dashboard
    """)
    total, compliant, non_compliant, overdue = rows[0]
    return jsonify({'total': total, 'compliant': compliant, 'non_compliant': non_compliant, 'overdue': overdue})


# ── Compliance dashboard ───────────────────────────────────────────────────
@app.route('/api/dashboard')
def dashboard():
    cols, rows = db.fetch_all("SELECT * FROM compliance_dashboard")
    return jsonify([dict(zip(cols, row)) for row in rows])


# ── Risk heatmap (FR21) ────────────────────────────────────────────────────
# Returns asset counts bucketed into a 5x5 risk grid for both inherent and
# residual risk. Used by the dashboard to render two side-by-side heatmaps.
# Score bands: 1-8 Low (green), 9-12 Medium (amber), 13-25 High (red).
@app.route('/api/risk-heatmap')
@login_required
def risk_heatmap():
    cols, rows = db.fetch_all("""
        SELECT asset_id, system_name,
               inherent_likelihood, inherent_impact,
               residual_likelihood, residual_impact
        FROM assets
        WHERE is_active = TRUE
    """)

    # Initialise empty 5x5 grids. Indexing: grid[likelihood-1][impact-1].
    # Each cell holds {'count': n, 'assets': [list of asset_ids]} so the
    # frontend can show a tooltip listing which assets fall into a cell.
    inherent = [[{'count': 0, 'assets': []} for _ in range(5)] for _ in range(5)]
    residual = [[{'count': 0, 'assets': []} for _ in range(5)] for _ in range(5)]
    classified = 0
    unclassified_assets = []

    for row in rows:
        record = dict(zip(cols, row))
        asset_id = record['asset_id']
        i_l = record.get('inherent_likelihood')
        i_i = record.get('inherent_impact')
        r_l = record.get('residual_likelihood')
        r_i = record.get('residual_impact')

        if all(v is not None and 1 <= v <= 5 for v in (i_l, i_i, r_l, r_i)):
            inherent[i_l - 1][i_i - 1]['count'] += 1
            inherent[i_l - 1][i_i - 1]['assets'].append(asset_id)
            residual[r_l - 1][r_i - 1]['count'] += 1
            residual[r_l - 1][r_i - 1]['assets'].append(asset_id)
            classified += 1
        else:
            unclassified_assets.append(asset_id)

    # Summary counts by risk band, derived from same source as the grids
    def band_summary(grid):
        low = sum(grid[l][i]['count'] for l in range(5) for i in range(5)
                  if 1 <= (l + 1) * (i + 1) <= 8)
        medium = sum(grid[l][i]['count'] for l in range(5) for i in range(5)
                     if 9 <= (l + 1) * (i + 1) <= 12)
        high = sum(grid[l][i]['count'] for l in range(5) for i in range(5)
                   if 13 <= (l + 1) * (i + 1) <= 25)
        return {'low': low, 'medium': medium, 'high': high}

    return jsonify({
        'inherent': inherent,
        'residual': residual,
        'inherent_summary': band_summary(inherent),
        'residual_summary': band_summary(residual),
        'classified': classified,
        'unclassified': len(unclassified_assets),
        'unclassified_assets': unclassified_assets
    })


# ── Risk Treatment Report (FR22) ───────────────────────────────────────────
# Generates an Excel report focused on the residual-risk profile of every
# asset. Sorted by residual score descending so the highest-risk items
# appear first. Used by Risk and Compliance teams to prioritise remediation.
@app.route('/api/report/risk-treatment')
@login_required
def risk_treatment_report():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime

    cols, rows = db.fetch_all("""
        SELECT asset_id, system_name, data_classification,
               inherent_likelihood, inherent_impact,
               (COALESCE(inherent_likelihood,0) * COALESCE(inherent_impact,0)) AS inherent_score,
               residual_likelihood, residual_impact,
               (COALESCE(residual_likelihood,0) * COALESCE(residual_impact,0)) AS residual_score,
               risk_treatment_plan, review_frequency, target_review_date,
               project_manager, key_risks
        FROM assets
        WHERE is_active = TRUE
        ORDER BY (COALESCE(residual_likelihood,0) * COALESCE(residual_impact,0)) DESC NULLS LAST,
                 asset_id ASC
    """)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Risk Treatment Detail'

    # Header row
    headers = [
        'Asset ID', 'System Name', 'Classification',
        'Inherent Likelihood', 'Inherent Impact', 'Inherent Score', 'Inherent Band',
        'Residual Likelihood', 'Residual Impact', 'Residual Score', 'Residual Band',
        'Risk Treatment Plan', 'Review Frequency', 'Target Review Date',
        'Project Manager', 'Key Risks'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F4F4F', end_color='4F4F4F', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Band classifier (matches the SQL view + heatmap)
    def band(score):
        if score is None or score == 0:
            return ''
        if score <= 8:
            return 'Low'
        if score <= 12:
            return 'Medium'
        return 'High'

    band_fills = {
        'Low':    PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'Medium': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'High':   PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }

    for r_idx, row in enumerate(rows, 2):
        record = dict(zip(cols, row))
        i_score = record.get('inherent_score') or 0
        r_score = record.get('residual_score') or 0
        i_band = band(i_score)
        r_band = band(r_score)

        values = [
            record.get('asset_id'),
            record.get('system_name'),
            record.get('data_classification'),
            record.get('inherent_likelihood'),
            record.get('inherent_impact'),
            i_score if i_score else '',
            i_band,
            record.get('residual_likelihood'),
            record.get('residual_impact'),
            r_score if r_score else '',
            r_band,
            record.get('risk_treatment_plan'),
            record.get('review_frequency'),
            record.get('target_review_date'),
            record.get('project_manager'),
            record.get('key_risks'),
        ]
        for c_idx, v in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

        # Colour the band cells
        if i_band:
            ws.cell(row=r_idx, column=7).fill = band_fills[i_band]
        if r_band:
            ws.cell(row=r_idx, column=11).fill = band_fills[r_band]

    # Column widths sized for readability
    # Positions 4-5 (Inherent Likelihood/Impact) and 8-9 (Residual Likelihood/Impact)
    # widened to fit the full header text after expanding from "L"/"I" abbreviations.
    widths = [12, 26, 14, 20, 17, 12, 12, 20, 17, 12, 12, 50, 14, 16, 24, 40]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    # ── Build the executive heatmap summary as a separate worksheet ──────────
    # Inserted at position 1 (the SECOND tab) — the detail sheet stays first
    # because the report's filename suggests "treatment list", and the summary
    # is supplementary context for review.
    summary = wb.create_sheet('Risk Heatmap Summary', 1)

    # Compute the 5x5 buckets by walking the rows we already fetched.
    # Each cell is just a count; we don't need the asset list here.
    inherent_grid = [[0] * 5 for _ in range(5)]  # [likelihood-1][impact-1]
    residual_grid = [[0] * 5 for _ in range(5)]
    classified_count = 0
    for row in rows:
        record = dict(zip(cols, row))
        i_l = record.get('inherent_likelihood')
        i_i = record.get('inherent_impact')
        r_l = record.get('residual_likelihood')
        r_i = record.get('residual_impact')
        if all(v is not None and 1 <= v <= 5 for v in (i_l, i_i, r_l, r_i)):
            inherent_grid[i_l - 1][i_i - 1] += 1
            residual_grid[r_l - 1][r_i - 1] += 1
            classified_count += 1
    unclassified_count = len(rows) - classified_count

    # ISO 31000 standard labels — match the dashboard for consistency
    likelihood_labels = ['Almost Certain', 'Likely', 'Possible', 'Unlikely', 'Rare']  # top→bottom
    impact_labels = ['Insignificant', 'Minor', 'Moderate', 'Major', 'Severe']         # left→right

    # Cell band classifier identical to detail sheet
    def cell_band(score):
        if score is None or score == 0:
            return None
        if score <= 8:
            return 'Low'
        if score <= 12:
            return 'Medium'
        return 'High'

    # ── Title row ────────────────────────────────────────────────────────────
    title_cell = summary.cell(row=1, column=1, value='Risk Heatmap Summary')
    title_cell.font = Font(bold=True, size=16, color='1F2937')
    summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    subtitle = summary.cell(
        row=2, column=1,
        value=f"Generated {datetime.now().strftime('%d/%m/%Y %H:%M')} · {classified_count} of {classified_count + unclassified_count} assets classified · ISO 31000 5×5 risk matrix"
    )
    subtitle.font = Font(italic=True, color='6B7280', size=10)
    summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    # ── Helper to render one 5x5 heatmap starting at a given row ─────────────
    section_header_font = Font(bold=True, size=12, color='1F2937')
    axis_font = Font(bold=True, size=10, color='4B5563')
    cell_count_font = Font(bold=True, size=11)
    legend_font = Font(size=10)

    def render_heatmap_block(start_row, title_text, grid):
        # Section title
        t = summary.cell(row=start_row, column=1, value=title_text)
        t.font = section_header_font
        summary.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)

        # Column headers (Impact axis)
        # Layout: col 1 = blank corner, col 2-6 = impact labels, col 7 = blank
        header_row = start_row + 2
        corner = summary.cell(row=header_row, column=1, value='Likelihood ↓ / Impact →')
        corner.font = Font(bold=True, italic=True, size=9, color='6B7280')
        corner.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for i, label in enumerate(impact_labels):
            c = summary.cell(row=header_row, column=2 + i, value=label)
            c.font = axis_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            c.border = thin_border

        # 5 grid rows — likelihood 5 (top, "Almost Certain") down to 1 ("Rare")
        for row_idx in range(5):
            l_value = 5 - row_idx  # 5, 4, 3, 2, 1
            grid_row = header_row + 1 + row_idx
            # Y-axis label
            y = summary.cell(row=grid_row, column=1, value=likelihood_labels[row_idx])
            y.font = axis_font
            y.alignment = Alignment(horizontal='center', vertical='center')
            y.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            y.border = thin_border
            # 5 impact cells
            for col_idx in range(5):
                i_value = col_idx + 1
                count = grid[l_value - 1][i_value - 1]
                score = l_value * i_value
                bnd = cell_band(score)
                cell = summary.cell(row=grid_row, column=2 + col_idx,
                                    value=count if count > 0 else None)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = cell_count_font
                cell.border = thin_border
                if bnd:
                    cell.fill = band_fills[bnd]

        # Band totals row
        totals_row = header_row + 6 + 1
        low = sum(grid[l][i] for l in range(5) for i in range(5) if 1 <= (l + 1) * (i + 1) <= 8)
        medium = sum(grid[l][i] for l in range(5) for i in range(5) if 9 <= (l + 1) * (i + 1) <= 12)
        high = sum(grid[l][i] for l in range(5) for i in range(5) if 13 <= (l + 1) * (i + 1) <= 25)
        totals_cell = summary.cell(
            row=totals_row, column=1,
            value=f"Band totals:    Low: {low}        Medium: {medium}        High: {high}"
        )
        totals_cell.font = Font(bold=True, size=10, color='374151')
        summary.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=7)

        # Return the next free row for the caller
        return totals_row + 2

    # Render Inherent block, then Residual block
    next_row = render_heatmap_block(4, 'INHERENT RISK (before controls)', inherent_grid)
    next_row = render_heatmap_block(next_row + 1, 'RESIDUAL RISK (after controls)', residual_grid)

    # ── Legend ──────────────────────────────────────────────────────────────
    legend_row = next_row + 1
    legend_title = summary.cell(row=legend_row, column=1, value='LEGEND')
    legend_title.font = section_header_font
    summary.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=7)

    legend_items = [
        ('Low (1–8)',    'Acceptable — operating within risk tolerance',                  'Low'),
        ('Medium (9–12)', 'Monitor — review controls at scheduled cadence',               'Medium'),
        ('High (13–25)', 'Escalate — treat or accept with documented justification',      'High'),
    ]
    for offset, (label, description, band_key) in enumerate(legend_items):
        r = legend_row + 2 + offset
        swatch = summary.cell(row=r, column=1, value='')
        swatch.fill = band_fills[band_key]
        swatch.border = thin_border
        l = summary.cell(row=r, column=2, value=label)
        l.font = Font(bold=True, size=10)
        d = summary.cell(row=r, column=3, value=description)
        d.font = legend_font
        summary.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

    if unclassified_count > 0:
        note_row = legend_row + 2 + len(legend_items) + 1
        n = summary.cell(
            row=note_row, column=1,
            value=f"Note: {unclassified_count} asset(s) are not classified and excluded from the heatmap counts."
        )
        n.font = Font(italic=True, size=9, color='B45309')
        summary.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)

    # Column widths for the summary sheet
    # Note: cannot use summary.cell(...).column_letter because row 1 col 1 is
    # part of a merged range (the title), and merged cells don't expose
    # column_letter. Use openpyxl's get_column_letter helper instead.
    from openpyxl.utils import get_column_letter
    summary_widths = [22, 14, 14, 14, 14, 14, 14]
    for col_idx, w in enumerate(summary_widths, 1):
        summary.column_dimensions[get_column_letter(col_idx)].width = w
    # Make the heatmap rows tall enough to read comfortably
    for r in range(6, 12):
        summary.row_dimensions[r].height = 26
    # Spacer + residual grid rows (positions shift by render_heatmap_block return)
    for r in range(13, 25):
        summary.row_dimensions[r].height = 26
    summary.row_dimensions[1].height = 28

    # Detail sheet stays as tab 1; Heatmap Summary is tab 2 (added above
    # via create_sheet(..., 1))

    # Export to memory + send
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Risk_Treatment_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    log_change(asset_id='SYSTEM', action='EXPORT', changed_by=request.current_user,
               notes='Risk Treatment Report generated')
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Asset register (with search and filter — FR11) ─────────────────────────
@app.route('/api/assets')
def assets():
    # Query parameters: q (free-text), classification, owner, compliance_status
    q              = request.args.get('q', '').strip()
    classification = request.args.get('classification', '').strip()
    owner          = request.args.get('owner', '').strip()
    compliance     = request.args.get('compliance_status', '').strip()

    # Build WHERE clauses dynamically
    conditions = ["a.is_active = TRUE"]
    params = []

    if q:
        conditions.append(
            "(a.asset_id ILIKE %s OR a.system_name ILIKE %s OR a.data_category ILIKE %s)"
        )
        like_q = f"%{q}%"
        params += [like_q, like_q, like_q]

    if classification:
        conditions.append("a.data_classification = %s")
        params.append(classification)

    if owner:
        conditions.append("a.data_owner ILIKE %s")
        params.append(f"%{owner}%")

    where = " AND ".join(conditions)

    query = f"""
        SELECT a.asset_id, a.system_name, a.data_category, a.data_classification,
               a.business_function, a.data_owner, a.custodian, a.data_location,
               a.who_has_access, a.access_mechanism, a.security_controls,
               a.encryption_in_transit, a.encryption_at_rest,
               a.retention_period, a.backup_retention, a.disposal_method,
               a.primary_use, a.policy_reference, a.control_mapping,
               a.last_review_date, a.remarks, a.description,
               a.data_sensitivity, a.business_criticality, a.risk_rating, a.key_risks,
               a.inherent_likelihood, a.inherent_impact,
               (COALESCE(a.inherent_likelihood,0) * COALESCE(a.inherent_impact,0)) AS inherent_score,
               a.residual_likelihood, a.residual_impact,
               (COALESCE(a.residual_likelihood,0) * COALESCE(a.residual_impact,0)) AS residual_score,
               a.risk_treatment_plan, a.review_frequency, a.target_review_date,
               a.project_manager, a.network_ports_protocols,
               a.is_active, {MANDATORY_SCORE_SQL} as mandatory_score
        FROM assets a
        WHERE {where}
        ORDER BY a.asset_id
    """

    cols, rows = db.fetch_all(query, params if params else None)

    result = []
    for row in rows:
        d = dict(zip(cols, [str(v) if v is not None else '' for v in row]))
        score = int(d.get('mandatory_score', 0))
        d['compliance_score'] = round((score / MANDATORY_FIELD_COUNT) * 100)
        d['compliance_status'] = 'Compliant' if score == MANDATORY_FIELD_COUNT else 'Non-Compliant'
        # Apply compliance_status filter if requested
        if compliance and d['compliance_status'] != compliance:
            continue
        result.append(d)
    return jsonify(result)


# ── CSV export of asset register (FR11 / Use Case) ─────────────────────────
@app.route('/api/assets/export/csv')
@login_required
def export_assets_csv():
    cols, rows = db.fetch_all("""
        SELECT asset_id, system_name, data_category, description,
               business_function, data_owner, custodian, data_classification, data_location,
               who_has_access, access_mechanism, security_controls,
               encryption_in_transit, encryption_at_rest, retention_period,
               backup_retention, disposal_method, primary_use,
               policy_reference, control_mapping, last_review_date, remarks,
               data_sensitivity, business_criticality, risk_rating, key_risks
        FROM assets WHERE is_active = TRUE ORDER BY asset_id
    """)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Asset ID', 'Asset Name', 'Data Category', 'Description',
        'Business Function', 'Data Owner', 'Data Custodian', 'Data Classification', 'Data Location',
        'Who Has Access', 'Access Mechanism', 'Security Controls',
        'Encryption in Transit', 'Encryption at Rest', 'Retention Period',
        'Backup Retention', 'Disposal Method', 'Permitted Use',
        'Policy Reference', 'Control Mapping', 'Last Review Date', 'Remarks',
        'Data Sensitivity', 'Business Criticality', 'Risk Rating', 'Key Risks'
    ])
    for row in rows:
        writer.writerow([str(v) if v is not None else '' for v in row])

    output.seek(0)
    filename = f"Asset_Register_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    log_change(asset_id='SYSTEM', action='EXPORT', changed_by=request.current_user,
               notes=f'CSV export downloaded: {filename}')
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv'
    )


# ── Category breakdown ─────────────────────────────────────────────────────
@app.route('/api/categories')
def categories():
    _, rows = db.fetch_all("SELECT data_category, COUNT(*) FROM assets WHERE is_active=TRUE GROUP BY data_category ORDER BY count DESC")
    return jsonify([{'category': r[0], 'count': r[1]} for r in rows])


# ── Classification breakdown ───────────────────────────────────────────────
@app.route('/api/classifications')
def classifications():
    _, rows = db.fetch_all("SELECT data_classification, COUNT(*) FROM assets WHERE is_active=TRUE GROUP BY data_classification ORDER BY count DESC")
    return jsonify([{'classification': r[0], 'count': r[1]} for r in rows])


# ── Activity feed ──────────────────────────────────────────────────────────
@app.route('/api/activity')
def activity():
    cols, rows = db.fetch_all("""
        SELECT changed_at, asset_id, action, field_changed, changed_by
        FROM audit_log
        WHERE asset_id IS NOT NULL
          AND asset_id != ''
          AND asset_id != 'SYSTEM'
          AND action IN ('INSERT','UPDATE','DELETE','VIEW','APPROVE','REJECT')
        ORDER BY changed_at DESC LIMIT 10
    """)
    return jsonify([dict(zip(cols, [str(v) if v is not None else '' for v in row])) for row in rows])


# ── Compliance trend ───────────────────────────────────────────────────────
@app.route('/api/trend')
def trend():
    _, rows = db.fetch_all("""
        SELECT DATE(run_at) as run_date,
               SUM(total_csv_records) as total,
               SUM(undocumented + misclassified + missing_fields) as findings
        FROM reconciliation_runs
        GROUP BY DATE(run_at)
        ORDER BY run_date ASC
        LIMIT 10
    """)
    return jsonify([{'date': str(r[0]), 'total': r[1], 'findings': r[2]} for r in rows])


# ── Asset history ──────────────────────────────────────────────────────────
@app.route('/api/assets/<asset_id>/history')
def asset_history(asset_id):
    cols, rows = db.fetch_all("""
        SELECT changed_at, action, field_changed, old_value, new_value, changed_by, notes
        FROM audit_log WHERE asset_id = %s ORDER BY changed_at DESC
    """, (asset_id,))
    return jsonify([dict(zip(cols, [str(v) if v is not None else '' for v in row])) for row in rows])


# ── Audit log ──────────────────────────────────────────────────────────────
@app.route('/api/audit')
def audit():
    # Cursor-based pagination: client passes ?before=<ISO_timestamp> to load
    # entries older than the last one shown. Default page size = 50.
    before = request.args.get('before')
    if before:
        cols, rows = db.fetch_all("""
            SELECT changed_at, asset_id, action, field_changed, old_value, new_value, changed_by
            FROM audit_log WHERE changed_at < %s ORDER BY changed_at DESC LIMIT 50
        """, (before,))
    else:
        cols, rows = db.fetch_all("""
            SELECT changed_at, asset_id, action, field_changed, old_value, new_value, changed_by
            FROM audit_log ORDER BY changed_at DESC LIMIT 50
        """)
    return jsonify([dict(zip(cols, [str(v) if v is not None else '' for v in row])) for row in rows])


# ── Findings ───────────────────────────────────────────────────────────────
@app.route('/api/findings')
def findings():
    cols, rows = db.fetch_all("""
        SELECT d.id, d.asset_id, d.finding_type, d.detail, d.csv_value, d.register_value,
               d.resolved, rr.run_at, rr.csv_filename
        FROM discrepancies d
        JOIN reconciliation_runs rr ON d.run_id = rr.id
        ORDER BY rr.run_at DESC LIMIT 50
    """)
    return jsonify([dict(zip(cols, [str(v) if v is not None else '' for v in row])) for row in rows])


# ── Resolve finding ────────────────────────────────────────────────────────
@app.route('/api/findings/<finding_id>/resolve', methods=['POST'])
@login_required
def resolve_finding(finding_id):
    db.execute("""
        UPDATE discrepancies
        SET resolved = TRUE, resolved_at = %s, resolved_by = %s
        WHERE id = %s
    """, (datetime.now(), request.current_user, finding_id))
    return jsonify({'success': True})


# ── Apply accepted findings (auto-patch assets) ────────────────────────────
@app.route('/api/findings/resolved', methods=['DELETE'])
@admin_required
def clear_resolved_findings():
    """Delete all resolved discrepancies to clean up the findings history."""
    cols, rows = db.fetch_all("SELECT COUNT(*) FROM discrepancies WHERE resolved = TRUE")
    count = rows[0][0] if rows else 0
    db.execute("DELETE FROM discrepancies WHERE resolved = TRUE")
    log_change(
        asset_id='SYSTEM', action='DELETE',
        changed_by=request.current_user,
        notes=f'Cleared {count} resolved finding(s) from history'
    )
    return jsonify({'success': True, 'cleared': count})


@app.route('/api/findings/apply', methods=['POST'])
@admin_required
def apply_findings():
    data = request.json
    accepted_ids = data.get('accepted', [])   # list of finding IDs to apply
    rejected_ids = data.get('rejected', [])   # list of finding IDs to reject (just resolve, no patch)

    applied = 0
    skipped = 0
    errors = []
    # Per-finding results so the frontend can navigate to the affected asset
    # after a single-finding Apply Fix. Only entries where status == 'applied'
    # and finding_type is editable (Misclassified / Missing Fields) drive the
    # modal-jump behaviour on the client side.
    results = []

    for finding_id in accepted_ids:
        try:
            cols, rows = db.fetch_all(
                "SELECT asset_id, finding_type, csv_value, register_value, detail, csv_row_data FROM discrepancies WHERE id = %s::uuid",
                (finding_id,)
            )
            if not rows:
                skipped += 1
                continue

            asset_id, finding_type, csv_value, register_value, detail, csv_row_data = rows[0]

            if finding_type == 'Misclassified':
                # Apply CSV classification to register
                valid = {'Public', 'Internal', 'Restricted', 'Confidential'}
                if csv_value in valid:
                    old_val = register_value
                    db.execute(
                        "UPDATE assets SET data_classification = %s WHERE asset_id = %s",
                        (csv_value, asset_id)
                    )
                    log_change(
                        asset_id=asset_id,
                        action='UPDATE',
                        field_changed='data_classification',
                        old_value=old_val,
                        new_value=csv_value,
                        changed_by=request.current_user,
                        notes=f'Auto-applied from reconciliation finding {finding_id}'
                    )
                    applied += 1
                    results.append({
                        'finding_id': finding_id,
                        'asset_id': asset_id,
                        'finding_type': finding_type,
                        'field_changed': 'data_classification',
                        'new_value': csv_value,
                        'editable': True,
                        'action': 'edit',
                        'status': 'applied'
                    })
                else:
                    errors.append(f'Finding {finding_id}: invalid classification value "{csv_value}"')
                    skipped += 1
                    continue

            elif finding_type == 'Missing Fields':
                # Extract field name from detail: "Mandatory field 'X' is empty"
                match = re.search(r"field '(.+?)' is empty", detail)
                if match:
                    field = match.group(1)
                    # Allow-list check — defence against SQL injection via
                    # tampered detail strings. The column name is interpolated
                    # directly into the UPDATE statement below because psycopg2
                    # cannot parameterise identifiers, so a whitelist is
                    # required.
                    if field not in UPDATABLE_ASSET_COLUMNS:
                        errors.append(f'Finding {finding_id}: field "{field}" is not an updatable asset column')
                        skipped += 1
                        continue
                    # Only patch if csv_value provides a value to fill
                    if csv_value and csv_value.strip():
                        db.execute(
                            f"UPDATE assets SET {field} = %s WHERE asset_id = %s",
                            (csv_value.strip(), asset_id)
                        )
                        log_change(
                            asset_id=asset_id,
                            action='UPDATE',
                            field_changed=field,
                            old_value='',
                            new_value=csv_value.strip(),
                            changed_by=request.current_user,
                            notes=f'Auto-applied from reconciliation finding {finding_id}'
                        )
                        applied += 1
                        results.append({
                            'finding_id': finding_id,
                            'asset_id': asset_id,
                            'finding_type': finding_type,
                            'field_changed': field,
                            'new_value': csv_value.strip(),
                            'editable': True,
                            'action': 'edit',
                            'status': 'applied'
                        })
                    else:
                        # CSV row exists but has no value for this missing field.
                        # We can't auto-apply, but we can acknowledge the finding
                        # and direct the user to fill the field manually. Frontend
                        # opens the Edit modal with the empty field highlighted.
                        db.execute("""
                            UPDATE discrepancies
                            SET resolved = TRUE, resolved_at = %s, resolved_by = %s
                            WHERE id = %s::uuid
                        """, (datetime.now(), request.current_user, finding_id))
                        applied += 1
                        results.append({
                            'finding_id': finding_id,
                            'asset_id': asset_id,
                            'finding_type': finding_type,
                            'field_changed': field,
                            'new_value': '',
                            'editable': True,
                            'action': 'edit_empty',
                            'status': 'applied'
                        })
                        continue
                else:
                    skipped += 1
                    continue

            elif finding_type == 'Undocumented':
                # Mark as resolved (acknowledged) even though we can not auto-create the asset
                db.execute("""
                    UPDATE discrepancies
                    SET resolved = TRUE, resolved_at = %s, resolved_by = %s
                    WHERE id = %s::uuid
                """, (datetime.now(), request.current_user, finding_id))
                applied += 1
                # Asset does not exist in the register. The frontend opens the
                # Add Asset modal pre-filled with whatever fields the CSV
                # provided so the user can confirm and complete the record.
                # csv_row_data is the original CSV row (JSONB on the
                # discrepancies table); it may be None on older findings
                # created before this column existed, in which case the
                # frontend falls back to pre-filling only the asset_id.
                prefill = {}
                if csv_row_data:
                    # psycopg2 returns JSONB as a dict already; older versions
                    # may return a str — handle both defensively.
                    if isinstance(csv_row_data, str):
                        try:
                            csv_row_data = json.loads(csv_row_data)
                        except Exception:
                            csv_row_data = {}
                    if isinstance(csv_row_data, dict):
                        # Only forward non-empty string values so we don't
                        # overwrite Add modal fields with blanks.
                        for k, v in csv_row_data.items():
                            if k and isinstance(v, (str, int, float)) and str(v).strip():
                                prefill[str(k).strip()] = str(v).strip()
                results.append({
                    'finding_id': finding_id,
                    'asset_id': asset_id,
                    'finding_type': finding_type,
                    'editable': True,
                    'action': 'create',
                    'prefill': prefill,
                    'status': 'applied'
                })
                continue

            # Mark finding as resolved
            db.execute("""
                UPDATE discrepancies
                SET resolved = TRUE, resolved_at = %s, resolved_by = %s
                WHERE id = %s::uuid
            """, (datetime.now(), request.current_user, finding_id))

        except Exception as e:
            errors.append(f'Finding {finding_id}: {str(e)}')
            skipped += 1

    # Mark rejected findings as resolved without patching
    for finding_id in rejected_ids:
        try:
            db.execute("""
                UPDATE discrepancies
                SET resolved = TRUE, resolved_at = %s, resolved_by = %s
                WHERE id = %s::uuid
            """, (datetime.now(), request.current_user, finding_id))
        except Exception as e:
            errors.append(f'Rejected finding {finding_id}: {str(e)}')

    return jsonify({
        'applied': applied,
        'skipped': skipped,
        'rejected': len(rejected_ids),
        'errors': errors,
        # Per-finding results enable the frontend to navigate to the asset
        # that was edited after a single-finding Apply Fix. Bulk applies
        # produce multiple entries; the frontend uses only the single-finding
        # case for the modal-jump UX.
        'results': results
    })


# ── Export findings PDF ────────────────────────────────────────────────────
@app.route('/api/findings/export/pdf')
def export_findings_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    cols, rows = db.fetch_all("""
        SELECT d.asset_id, d.finding_type, d.detail, d.csv_value, d.register_value,
               d.resolved, rr.run_at, rr.csv_filename
        FROM discrepancies d
        JOIN reconciliation_runs rr ON d.run_id = rr.id
        ORDER BY rr.run_at DESC LIMIT 100
    """)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', fontSize=16, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1F3864'), spaceAfter=6)
    sub_style = ParagraphStyle('sub', fontSize=10, fontName='Helvetica',
                                textColor=colors.HexColor('#475569'), spaceAfter=20)
    elements = []
    elements.append(Paragraph("Reconciliation Findings Report", title_style))
    elements.append(Paragraph(
        f"Eskan Bank — Information Asset Registry System | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
        sub_style
    ))

    total = len(rows)
    undoc = sum(1 for r in rows if r[1] == 'Undocumented')
    miscls = sum(1 for r in rows if r[1] == 'Misclassified')
    missing = sum(1 for r in rows if r[1] == 'Missing Fields')
    resolved = sum(1 for r in rows if r[5])

    summary_data = [['Total', 'Undocumented', 'Misclassified', 'Missing Fields', 'Resolved'],
                    [str(total), str(undoc), str(miscls), str(missing), str(resolved)]]
    summary_table = Table(summary_data, colWidths=[5*cm]*5)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#D6E4F0')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFBFBF')),
        ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*cm))

    table_data = [['Run Date', 'Asset ID', 'Type', 'Detail', 'CSV Value', 'Register Value', 'Status']]
    type_colors = {'Undocumented': '#FFDAD6', 'Misclassified': '#FFF2CC', 'Missing Fields': '#D6E4F0'}
    for row in rows:
        table_data.append([
            str(row[6])[:10] if row[6] else '',
            str(row[0] or ''),
            str(row[1] or ''),
            str(row[2] or '')[:60],
            str(row[3] or '')[:20],
            str(row[4] or '')[:20],
            'Resolved' if row[5] else 'Open'
        ])

    ft = Table(table_data, colWidths=[3*cm, 2.5*cm, 3*cm, 9*cm, 3.5*cm, 3.5*cm, 2.5*cm], repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F3864')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#BFBFBF')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5)
    ]
    for i, row in enumerate(rows, start=1):
        bg = type_colors.get(str(row[1]), '#FFFFFF')
        style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor(bg)))
    ft.setStyle(TableStyle(style))
    elements.append(ft)
    doc.build(elements)
    buffer.seek(0)
    filename = f"Findings_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


# ── Generate Excel report ──────────────────────────────────────────────────
@app.route('/api/report', methods=['POST'])
@login_required
def generate_report():
    from report import generate_report as gen
    path = gen(output_dir='.')
    log_change(asset_id='SYSTEM', action='EXPORT', changed_by=request.current_user,
               notes='Excel report generated')
    return send_file(path, as_attachment=True,
                     download_name=os.path.basename(path),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Run reconciliation ─────────────────────────────────────────────────────
@app.route('/api/reconcile', methods=['POST'])
@login_required
def reconcile():
    from reconciliation import run_reconciliation
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    filename = (f.filename or '').lower()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = f'uploaded_{timestamp}.csv'

    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Convert Excel to CSV before passing to the reconciliation engine
            # (which only reads CSV). We use openpyxl directly to avoid adding
            # a pandas dependency. Only the first worksheet is read; further
            # sheets are ignored, matching the standard one-sheet-per-upload
            # convention used by Bahraini regulatory templates.
            from openpyxl import load_workbook
            xlsx_path = f'uploaded_{timestamp}.xlsx'
            f.save(xlsx_path)
            try:
                wb = load_workbook(xlsx_path, data_only=True, read_only=True)
                ws = wb.active
                with open(csv_path, 'w', newline='', encoding='utf-8') as out:
                    writer = csv.writer(out)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(['' if v is None else str(v) for v in row])
                wb.close()
            finally:
                if os.path.exists(xlsx_path):
                    os.remove(xlsx_path)
        else:
            f.save(csv_path)

        run_id, findings = run_reconciliation(csv_path, run_by=request.current_user)
        return jsonify({
            'run_id': str(run_id),
            'total_findings': len(findings),
            'undocumented': sum(1 for x in findings if x['finding_type'] == 'Undocumented'),
            'misclassified': sum(1 for x in findings if x['finding_type'] == 'Misclassified'),
            'missing_fields': sum(1 for x in findings if x['finding_type'] == 'Missing Fields'),
        })
    except Exception as e:
        return jsonify({'error': f'Could not process file: {str(e)}'}), 400
    finally:
        if os.path.exists(csv_path):
            os.remove(csv_path)


# ── CSV / XLSX Bulk Import ─────────────────────────────────────────────────
@app.route('/api/import', methods=['POST'])
@login_required
def bulk_import():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    filename = (f.filename or '').lower()

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Read Excel into a list of dict rows using openpyxl directly. The first
        # row is treated as the header. Empty cells become empty strings to
        # match csv.DictReader behaviour and keep downstream validation logic
        # unchanged. Only the first worksheet is read.
        from openpyxl import load_workbook
        try:
            wb = load_workbook(f, data_only=True, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                return jsonify({'error': 'Empty workbook'}), 400
            header = [('' if h is None else str(h).strip()) for h in header]
            rows_data = []
            for r in rows_iter:
                row_dict = {}
                for i, cell in enumerate(r):
                    if i < len(header) and header[i]:
                        row_dict[header[i]] = '' if cell is None else str(cell)
                if any(v.strip() for v in row_dict.values() if isinstance(v, str)):
                    rows_data.append(row_dict)
            wb.close()
            reader = iter(rows_data)
        except Exception as e:
            return jsonify({'error': f'Could not read Excel file: {str(e)}'}), 400
    else:
        content = f.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))

    required_fields = ['asset_id', 'system_name', 'data_category', 'data_classification', 'data_owner']
    valid_classifications = {'Public', 'Internal', 'Restricted', 'Confidential'}

    # ── Non-admin: submit entire import to pending queue ──────────────────
    if request.current_role != 'admin':
        rows_list = list(reader)
        submitted = 0
        skipped = 0
        errors = []
        for i, row in enumerate(rows_list, start=2):
            missing = [field for field in required_fields if not row.get(field, '').strip()]
            if missing:
                errors.append(f"Row {i}: missing {', '.join(missing)}")
                skipped += 1
                continue
            classification = row.get('data_classification', '').strip()
            if classification not in valid_classifications:
                errors.append(f"Row {i}: invalid data_classification '{classification}'")
                skipped += 1
                continue
            db.execute("""
                INSERT INTO pending_changes (asset_id, action, proposed_data, submitted_by)
                VALUES (%s, 'INSERT', %s, %s)
            """, (row.get('asset_id', '').strip(), json.dumps({k: v.strip() for k, v in row.items()}), request.current_user))
            submitted += 1
        if submitted > 0:
            _, admin_rows = db.fetch_all(
                "SELECT email FROM users WHERE role = 'admin' AND is_active = TRUE AND email IS NOT NULL"
            )
            for admin_row in admin_rows:
                if admin_row[0]:
                    send_email_async(send_pending_change_notification, 
                        admin_row[0], request.current_user, 'BULK IMPORT', f'{submitted} records'
                    )
        return jsonify({'inserted': 0, 'skipped': skipped, 'errors': errors[:10],
                        'pending': True, 'submitted': submitted,
                        'message': f'{submitted} records submitted for administrator approval'})

    # ── Admin: apply directly ─────────────────────────────────────────────
    inserted = 0
    skipped = 0
    errors = []

    for i, row in enumerate(reader, start=2):
        missing = [field for field in required_fields if not row.get(field, '').strip()]
        if missing:
            errors.append(f"Row {i}: missing {', '.join(missing)}")
            skipped += 1
            continue

        classification = row.get('data_classification', '').strip()
        if classification not in valid_classifications:
            errors.append(f"Row {i}: invalid data_classification '{classification}'")
            skipped += 1
            continue

        try:
            db.execute("""
                INSERT INTO assets (
                    asset_id, system_name, data_category, description,
                    business_function, data_owner, custodian, data_classification, data_location,
                    who_has_access, access_mechanism, security_controls,
                    encryption_in_transit, encryption_at_rest,
                    retention_period, backup_retention, disposal_method,
                    primary_use, policy_reference, control_mapping,
                    last_review_date, remarks,
                    data_sensitivity, business_criticality, risk_rating, key_risks,
                    created_by
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (asset_id) DO UPDATE SET
                    system_name = EXCLUDED.system_name,
                    data_category = EXCLUDED.data_category,
                    data_classification = EXCLUDED.data_classification,
                    data_owner = EXCLUDED.data_owner,
                    business_function = EXCLUDED.business_function,
                    data_sensitivity = EXCLUDED.data_sensitivity,
                    business_criticality = EXCLUDED.business_criticality,
                    risk_rating = EXCLUDED.risk_rating,
                    key_risks = EXCLUDED.key_risks,
                    updated_at = NOW()
            """, (
                row.get('asset_id', '').strip(),
                row.get('system_name', '').strip(),
                row.get('data_category', '').strip(),
                row.get('description', '').strip() or None,
                row.get('business_function', '').strip() or None,
                row.get('data_owner', '').strip(),
                row.get('custodian', '').strip() or None,
                classification,
                row.get('data_location', '').strip() or None,
                row.get('who_has_access', '').strip() or None,
                row.get('access_mechanism', '').strip() or None,
                row.get('security_controls', '').strip() or None,
                row.get('encryption_in_transit', '').strip() or None,
                row.get('encryption_at_rest', '').strip() or None,
                row.get('retention_period', '').strip() or None,
                row.get('backup_retention', '').strip() or None,
                row.get('disposal_method', '').strip() or None,
                row.get('primary_use', '').strip() or None,
                row.get('policy_reference', '').strip() or None,
                row.get('control_mapping', '').strip() or None,
                row.get('last_review_date', '').strip() or None,
                row.get('remarks', '').strip() or None,
                row.get('data_sensitivity', '').strip() or None,
                row.get('business_criticality', '').strip() or None,
                row.get('risk_rating', '').strip() or None,
                row.get('key_risks', '').strip() or None,
                request.current_user
            ))
            log_change(
                asset_id=row.get('asset_id', '').strip(),
                action='INSERT',
                changed_by=request.current_user,
                notes=f"Bulk import from {f.filename}"
            )
            inserted += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")
            skipped += 1

    return jsonify({'inserted': inserted, 'skipped': skipped, 'errors': errors[:10]})


# ── Download import template ───────────────────────────────────────────────
@app.route('/api/import/template')
def import_template():
    # Allowed values for constrained columns (single source of truth).
    # Matches the dropdown options in the Add/Edit Asset form so any bulk
    # import is validated identically. Numeric scales are 1-5 (likelihood
    # and impact). risk_rating remains for backward compatibility but is
    # deprecated in the form in favour of the computed inherent/residual
    # bands.
    allowed_values = {
        'data_category':         ['Financial Data', 'Personal Data', 'Regulatory Data', 'Internal Operations', 'Internal Communications', 'Operational / Log Data'],
        'data_classification':   ['Public', 'Internal', 'Restricted', 'Confidential'],
        'encryption_in_transit': ['Yes', 'No', 'Partial'],
        'encryption_at_rest':    ['Yes', 'No', 'Partial'],
        'business_criticality':  ['Critical', 'High', 'Medium', 'Low'],
        'risk_rating':           ['High', 'Medium', 'Low'],
        'inherent_likelihood':   ['1', '2', '3', '4', '5'],
        'inherent_impact':       ['1', '2', '3', '4', '5'],
        'residual_likelihood':   ['1', '2', '3', '4', '5'],
        'residual_impact':       ['1', '2', '3', '4', '5'],
        'review_frequency':      ['Weekly', 'Monthly', 'Quarterly', 'Semi-Annual', 'Annual'],
    }

    headers = [
        'asset_id', 'system_name', 'data_category', 'description',
        'business_function', 'data_owner', 'custodian', 'data_classification', 'data_location',
        'who_has_access', 'access_mechanism', 'security_controls',
        'encryption_in_transit', 'encryption_at_rest',
        'retention_period', 'backup_retention', 'disposal_method', 'primary_use',
        'policy_reference', 'control_mapping', 'last_review_date', 'remarks',
        'data_sensitivity', 'business_criticality', 'risk_rating', 'key_risks',
        'inherent_likelihood', 'inherent_impact',
        'residual_likelihood', 'residual_impact',
        'risk_treatment_plan', 'review_frequency', 'target_review_date',
        'project_manager', 'network_ports_protocols'
    ]
    example_row = [
        'AR-CBS-001', 'Core Banking System', 'Financial Data',
        'Primary banking platform holding customer accounts and transactions',
        'Retail Banking', 'Head of Retail Banking', 'IT Operations',
        'Restricted', 'On-Premises Data Centre', 'Retail Banking Staff',
        'RBAC with SSO', 'MFA, Firewall, IDS', 'Yes', 'Yes',
        '7 Years', '1 Year', 'Secure Cryptographic Erasure',
        'Account management and transaction processing',
        'IS-POL-001', 'CBB OM-5.5, ISO 27001 A.5.9, A.8.12',
        '2025-01-01', 'Reviewed annually',
        'High (PII + Financial Data)', 'Critical', 'High',
        'Unauthorised access, data leakage, regulatory breach',
        '4', '5',
        '2', '3',
        'MFA, RBAC, encryption at rest and in transit, SIEM monitoring, quarterly access reviews',
        'Quarterly', '2025-04-01',
        'AGM Enterprise Digital Optimization', 'TLS 1.3 / port 443 (HTTPS); PostgreSQL 5432'
    ]

    fmt = (request.args.get('format', 'csv') or 'csv').lower()

    if fmt == 'xlsx':
        # Excel template with bold headers, freeze pane, data validation
        # dropdowns on constrained columns, hover comments listing allowed
        # values, and a separate Reference sheet showing all valid values.
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Asset Import'
        ws.append(headers)
        ws.append(example_row)

        header_font = Font(bold=True, color='FFFFFF', name='Arial')
        header_fill = PatternFill('solid', start_color='1F2937')
        constrained_fill = PatternFill('solid', start_color='2563EB')  # blue tint for constrained columns
        for col_idx, header_name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            # Highlight constrained columns so the user knows where dropdowns are
            cell.fill = constrained_fill if header_name in allowed_values else header_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.column_dimensions[cell.column_letter].width = 22

            # Add a hover comment listing allowed values
            if header_name in allowed_values:
                opts = ', '.join(allowed_values[header_name])
                cell.comment = Comment(
                    f"Allowed values:\n{opts}",
                    "System"
                )
                cell.comment.width = 320
                cell.comment.height = 80

        # Add Excel data validation (in-cell dropdowns) for constrained
        # columns. Applied from row 2 down to row 1000 so users can paste
        # large batches and still get validation hints.
        for col_idx, header_name in enumerate(headers, start=1):
            if header_name not in allowed_values:
                continue
            opts = allowed_values[header_name]
            # Excel inline list validations are capped at 255 chars; ours
            # are all well under so we can embed directly.
            formula = '"' + ','.join(opts) + '"'
            dv = DataValidation(
                type='list',
                formula1=formula,
                allow_blank=True,
                showDropDown=False,  # show the dropdown arrow
                showErrorMessage=True,
                errorTitle='Invalid value',
                error=f"Must be one of: {', '.join(opts)}",
                promptTitle=f"{header_name}",
                prompt=f"Allowed: {', '.join(opts)}",
                showInputMessage=True,
            )
            col_letter = get_column_letter(col_idx)
            dv.add(f"{col_letter}2:{col_letter}1000")
            ws.add_data_validation(dv)

        ws.freeze_panes = 'A2'

        # Reference sheet — one column per constrained field, listing every
        # allowed value vertically so users can read or copy from it.
        ref = wb.create_sheet('Allowed Values')
        ref_headers = list(allowed_values.keys())
        ref.append(ref_headers)
        max_rows = max(len(v) for v in allowed_values.values())
        for r in range(max_rows):
            row_values = []
            for h in ref_headers:
                vals = allowed_values[h]
                row_values.append(vals[r] if r < len(vals) else '')
            ref.append(row_values)
        for col_idx, h in enumerate(ref_headers, start=1):
            cell = ref.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = constrained_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ref.column_dimensions[cell.column_letter].width = 22
        ref.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf,
                         as_attachment=True,
                         download_name='asset_register_import_template.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Default: CSV. csv.DictReader has no concept of comments or validation,
    # so we keep the file structurally simple: header row, one example row.
    # Valid values are documented in the Import Guide panel in the UI and on
    # the Allowed Values sheet of the XLSX template.
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(example_row)
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()),
                     as_attachment=True, download_name='asset_register_import_template.csv',
                     mimetype='text/csv')


# ── Add asset ──────────────────────────────────────────────────────────────
@app.route('/api/assets', methods=['POST'])
@login_required
def add_asset():
    data = request.json
    classification = data.get('data_classification', '')
    valid_classifications = {'Public', 'Internal', 'Restricted', 'Confidential'}
    if classification not in valid_classifications:
        return jsonify({'error': f"Invalid data_classification. Must be one of: {', '.join(sorted(valid_classifications))}"}), 400

    # Non-admin: submit to pending queue
    if request.current_role != 'admin':
        print(f"  [PENDING] INSERT submitted by {request.current_user} for asset {data.get('asset_id')}")
        db.execute("""
            INSERT INTO pending_changes (asset_id, action, proposed_data, submitted_by)
            VALUES (%s, 'INSERT', %s, %s)
        """, (data.get('asset_id'), json.dumps(data), request.current_user))
        # Notify admins
        _, admin_rows = db.fetch_all(
            "SELECT email FROM users WHERE role = 'admin' AND is_active = TRUE AND email IS NOT NULL"
        )
        for row in admin_rows:
            if row[0]:
                send_email_async(send_pending_change_notification, 
                    row[0], request.current_user, 'INSERT', data.get('asset_id', 'NEW')
                )
        return jsonify({'success': True, 'pending': True,
                        'message': 'Change submitted for administrator approval'})

    # Admin role: apply directly
    try:
        db.execute("""
            INSERT INTO assets (
                asset_id, system_name, data_category, description,
                business_function, data_owner, custodian, data_classification, data_location,
                who_has_access, access_mechanism, security_controls,
                encryption_in_transit, encryption_at_rest,
                retention_period, backup_retention, disposal_method,
                primary_use, policy_reference, control_mapping,
                last_review_date, remarks,
                data_sensitivity, business_criticality, risk_rating, key_risks,
                inherent_likelihood, inherent_impact, residual_likelihood, residual_impact,
                risk_treatment_plan, review_frequency, target_review_date,
                project_manager, network_ports_protocols,
                created_by
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            clean(data.get('asset_id')), clean(data.get('system_name')),
            clean(data.get('data_category')), clean(data.get('description')),
            clean(data.get('business_function')), clean(data.get('data_owner')),
            clean(data.get('custodian')), classification,
            clean(data.get('data_location')), clean(data.get('who_has_access')),
            clean(data.get('access_mechanism')), clean(data.get('security_controls')),
            clean(data.get('encryption_in_transit')), clean(data.get('encryption_at_rest')),
            clean(data.get('retention_period')), clean(data.get('backup_retention')),
            clean(data.get('disposal_method')), clean(data.get('primary_use')),
            clean(data.get('policy_reference')), clean(data.get('control_mapping')),
            clean(data.get('last_review_date')), clean(data.get('remarks')),
            clean(data.get('data_sensitivity')), clean(data.get('business_criticality')),
            clean(data.get('risk_rating')), clean(data.get('key_risks')),
            clean_int(data.get('inherent_likelihood'), 1, 5),
            clean_int(data.get('inherent_impact'), 1, 5),
            clean_int(data.get('residual_likelihood'), 1, 5),
            clean_int(data.get('residual_impact'), 1, 5),
            clean(data.get('risk_treatment_plan')),
            clean(data.get('review_frequency')),
            clean(data.get('target_review_date')),
            clean(data.get('project_manager')),
            clean(data.get('network_ports_protocols')),
            request.current_user
        ))
    except psycopg2.errors.UniqueViolation:
        # An asset with this ID already exists. Check whether it's active or
        # soft-deleted: if it was deleted previously, treat the re-add as a
        # restore (resurrect the row with the new field values). If the row
        # is active, refuse the insert as before.
        asset_id_val = clean(data.get('asset_id'))
        _, existing = db.fetch_all(
            "SELECT is_active FROM assets WHERE asset_id = %s",
            (asset_id_val,)
        )
        if not existing:
            # Shouldn't happen — UniqueViolation but no row found. Re-raise.
            return jsonify({'error': f"Constraint violation on asset_id '{asset_id_val}'."}), 409
        is_active_existing = existing[0][0]
        if is_active_existing:
            # Active duplicate — keep the original behaviour.
            return jsonify({
                'error': f"An asset with ID '{asset_id_val}' already exists in the register. "
                         f"To update it, search for it in the Asset Register and use the Edit button instead."
            }), 409
        # Soft-deleted duplicate — restore by overwriting all fields with the
        # new data and re-activating. This preserves the audit history (the
        # original INSERT and DELETE entries remain in audit_log) and lets the
        # user re-add an asset that they previously removed.
        db.execute("""
            UPDATE assets SET
                system_name = %s, data_category = %s, description = %s,
                business_function = %s, data_owner = %s, custodian = %s,
                data_classification = %s, data_location = %s, who_has_access = %s,
                access_mechanism = %s, security_controls = %s,
                encryption_in_transit = %s, encryption_at_rest = %s,
                retention_period = %s, backup_retention = %s, disposal_method = %s,
                primary_use = %s, policy_reference = %s, control_mapping = %s,
                last_review_date = %s, remarks = %s,
                data_sensitivity = %s, business_criticality = %s,
                risk_rating = %s, key_risks = %s,
                inherent_likelihood = %s, inherent_impact = %s,
                residual_likelihood = %s, residual_impact = %s,
                risk_treatment_plan = %s, review_frequency = %s,
                target_review_date = %s, project_manager = %s,
                network_ports_protocols = %s,
                is_active = TRUE,
                created_by = %s
            WHERE asset_id = %s
        """, (
            clean(data.get('system_name')),
            clean(data.get('data_category')), clean(data.get('description')),
            clean(data.get('business_function')), clean(data.get('data_owner')),
            clean(data.get('custodian')), classification,
            clean(data.get('data_location')), clean(data.get('who_has_access')),
            clean(data.get('access_mechanism')), clean(data.get('security_controls')),
            clean(data.get('encryption_in_transit')), clean(data.get('encryption_at_rest')),
            clean(data.get('retention_period')), clean(data.get('backup_retention')),
            clean(data.get('disposal_method')), clean(data.get('primary_use')),
            clean(data.get('policy_reference')), clean(data.get('control_mapping')),
            clean(data.get('last_review_date')), clean(data.get('remarks')),
            clean(data.get('data_sensitivity')), clean(data.get('business_criticality')),
            clean(data.get('risk_rating')), clean(data.get('key_risks')),
            clean_int(data.get('inherent_likelihood'), 1, 5),
            clean_int(data.get('inherent_impact'), 1, 5),
            clean_int(data.get('residual_likelihood'), 1, 5),
            clean_int(data.get('residual_impact'), 1, 5),
            clean(data.get('risk_treatment_plan')),
            clean(data.get('review_frequency')),
            clean(data.get('target_review_date')),
            clean(data.get('project_manager')),
            clean(data.get('network_ports_protocols')),
            request.current_user,
            asset_id_val
        ))
        log_change(
            asset_id=asset_id_val,
            action='INSERT',
            changed_by=request.current_user,
            notes='Asset restored from soft-deleted state'
        )
        return jsonify({'success': True, 'restored': True,
                        'message': f"Asset '{asset_id_val}' was previously deleted and has been restored with the new data."})
    except psycopg2.IntegrityError as e:
        return jsonify({'error': f"Database constraint violation: {str(e).splitlines()[0]}"}), 400

    log_change(
        asset_id=data.get('asset_id'),
        action='INSERT',
        changed_by=request.current_user,
        notes='Asset created via web interface'
    )
    return jsonify({'success': True})


# ── Edit asset ─────────────────────────────────────────────────────────────
@app.route('/api/assets/<asset_id>', methods=['PUT'])
@login_required
def edit_asset(asset_id):
    data = request.json
    classification = data.get('data_classification', '')
    valid_classifications = {'Public', 'Internal', 'Restricted', 'Confidential'}
    if classification and classification not in valid_classifications:
        return jsonify({'error': f"Invalid data_classification. Must be one of: {', '.join(sorted(valid_classifications))}"}), 400

    # Non-admin: submit each changed field to pending queue
    if request.current_role != 'admin':
        print(f"  [PENDING] UPDATE submitted by {request.current_user} for asset {asset_id}")
        cols, rows = db.fetch_all("SELECT * FROM assets WHERE asset_id = %s", (asset_id,))
        old_record = dict(zip(cols, rows[0])) if rows else {}
        audit_fields = [
            'system_name', 'data_category', 'description', 'business_function',
            'data_owner', 'custodian',
            'data_classification', 'data_location', 'who_has_access', 'access_mechanism',
            'security_controls', 'encryption_in_transit', 'encryption_at_rest',
            'retention_period', 'backup_retention', 'disposal_method',
            'primary_use', 'policy_reference', 'control_mapping', 'last_review_date', 'remarks',
            'data_sensitivity', 'business_criticality', 'risk_rating', 'key_risks',
            'inherent_likelihood', 'inherent_impact', 'residual_likelihood', 'residual_impact',
            'risk_treatment_plan', 'review_frequency', 'target_review_date',
            'project_manager', 'network_ports_protocols'
        ]
        changes_submitted = 0
        for field in audit_fields:
            new_val = data.get(field)
            old_val = str(old_record.get(field, '')) if old_record.get(field) is not None else ''
            new_val_str = str(new_val) if new_val is not None else ''
            if old_val != new_val_str:
                db.execute("""
                    INSERT INTO pending_changes
                    (asset_id, action, field_changed, old_value, new_value, submitted_by)
                    VALUES (%s, 'UPDATE', %s, %s, %s, %s)
                """, (asset_id, field, old_val, new_val_str, request.current_user))
                changes_submitted += 1

        if changes_submitted > 0:
            _, admin_rows = db.fetch_all(
                "SELECT email FROM users WHERE role = 'admin' AND is_active = TRUE AND email IS NOT NULL"
            )
            for row in admin_rows:
                if row[0]:
                    send_email_async(send_pending_change_notification, 
                        row[0], request.current_user, 'UPDATE', asset_id
                    )
        return jsonify({'success': True, 'pending': True,
                        'message': f'{changes_submitted} change(s) submitted for administrator approval'})

    # Admin role: apply directly
    cols, rows = db.fetch_all("SELECT * FROM assets WHERE asset_id = %s", (asset_id,))
    old_record = dict(zip(cols, rows[0])) if rows else {}

    db.execute("""
        UPDATE assets SET system_name=%s, data_category=%s, description=%s,
            business_function=%s, data_owner=%s, custodian=%s, data_classification=%s,
            data_location=%s, who_has_access=%s, access_mechanism=%s,
            security_controls=%s, encryption_in_transit=%s, encryption_at_rest=%s,
            retention_period=%s, backup_retention=%s, disposal_method=%s,
            primary_use=%s, policy_reference=%s, control_mapping=%s,
            last_review_date=%s, remarks=%s,
            data_sensitivity=%s, business_criticality=%s, risk_rating=%s, key_risks=%s,
            inherent_likelihood=%s, inherent_impact=%s,
            residual_likelihood=%s, residual_impact=%s,
            risk_treatment_plan=%s, review_frequency=%s, target_review_date=%s,
            project_manager=%s, network_ports_protocols=%s
        WHERE asset_id=%s
    """, (
        clean(data.get('system_name')), clean(data.get('data_category')),
        clean(data.get('description')), clean(data.get('business_function')),
        clean(data.get('data_owner')), clean(data.get('custodian')),
        classification or clean(data.get('data_classification')),
        clean(data.get('data_location')), clean(data.get('who_has_access')),
        clean(data.get('access_mechanism')), clean(data.get('security_controls')),
        clean(data.get('encryption_in_transit')), clean(data.get('encryption_at_rest')),
        clean(data.get('retention_period')), clean(data.get('backup_retention')),
        clean(data.get('disposal_method')), clean(data.get('primary_use')),
        clean(data.get('policy_reference')), clean(data.get('control_mapping')),
        clean(data.get('last_review_date')), clean(data.get('remarks')),
        clean(data.get('data_sensitivity')), clean(data.get('business_criticality')),
        clean(data.get('risk_rating')), clean(data.get('key_risks')),
        clean_int(data.get('inherent_likelihood'), 1, 5),
        clean_int(data.get('inherent_impact'), 1, 5),
        clean_int(data.get('residual_likelihood'), 1, 5),
        clean_int(data.get('residual_impact'), 1, 5),
        clean(data.get('risk_treatment_plan')),
        clean(data.get('review_frequency')),
        clean(data.get('target_review_date')),
        clean(data.get('project_manager')),
        clean(data.get('network_ports_protocols')),
        asset_id
    ))

    audit_fields = [
        'system_name', 'data_category', 'description', 'business_function',
        'data_owner', 'custodian', 'data_classification', 'data_location',
        'who_has_access', 'access_mechanism', 'security_controls',
        'encryption_in_transit', 'encryption_at_rest',
        'retention_period', 'backup_retention', 'disposal_method',
        'primary_use', 'policy_reference', 'control_mapping', 'last_review_date',
        'remarks', 'data_sensitivity', 'business_criticality', 'risk_rating', 'key_risks',
        'inherent_likelihood', 'inherent_impact', 'residual_likelihood', 'residual_impact',
        'risk_treatment_plan', 'review_frequency', 'target_review_date',
        'project_manager', 'network_ports_protocols'
    ]
    for field in audit_fields:
        new_val = data.get(field)
        old_val = str(old_record.get(field, '')) if old_record.get(field) is not None else ''
        new_val_str = str(new_val) if new_val is not None else ''
        if old_val != new_val_str:
            log_change(
                asset_id=asset_id,
                action='UPDATE',
                field_changed=field,
                old_value=old_val,
                new_value=new_val_str,
                changed_by=request.current_user
            )

    return jsonify({'success': True})


# ── Get single asset ───────────────────────────────────────────────────────
@app.route('/api/assets/<asset_id>', methods=['GET'])
@login_required
def get_asset(asset_id):
    cols, rows = db.fetch_all("SELECT * FROM assets WHERE asset_id = %s", (asset_id,))
    if not rows:
        return jsonify({'error': 'Not found'}), 404
    # Log view event for audit trail
    log_change(asset_id=asset_id, action='VIEW',
               changed_by=request.current_user,
               notes=f'Asset viewed from {request.remote_addr}')
    return jsonify(dict(zip(cols, [str(v) if v is not None else '' for v in rows[0]])))


# ── Deactivate asset (soft delete) ─────────────────────────────────────────
@app.route('/api/assets/<asset_id>', methods=['DELETE'])
@login_required
def deactivate_asset(asset_id):
    # Non-admin: submit to pending queue
    if request.current_role != 'admin':
        db.execute("""
            INSERT INTO pending_changes (asset_id, action, submitted_by)
            VALUES (%s, 'DELETE', %s)
        """, (asset_id, request.current_user))
        _, admin_rows = db.fetch_all(
            "SELECT email FROM users WHERE role = 'admin' AND is_active = TRUE AND email IS NOT NULL"
        )
        for row in admin_rows:
            if row[0]:
                send_email_async(send_pending_change_notification, row[0], request.current_user, 'DELETE', asset_id)
        return jsonify({'success': True, 'pending': True,
                        'message': 'Delete request submitted for administrator approval'})
    # Admin: apply directly
    db.execute("UPDATE assets SET is_active = FALSE WHERE asset_id = %s", (asset_id,))
    log_change(asset_id=asset_id, action='DELETE', changed_by=request.current_user,
               notes='Asset deactivated (soft delete)')
    return jsonify({'success': True})


# ── User Management ───────────────────────────────────────────────────────
@app.route('/api/users', methods=['GET'])
@admin_required
def list_users():
    cols, rows = db.fetch_all("""
        SELECT user_id, username, full_name, email, role, is_active, is_approved, last_login, created_at
        FROM users ORDER BY created_at DESC
    """)
    return jsonify([dict(zip(cols, [str(v) if v is not None else '' for v in row])) for row in rows])


@app.route('/api/users/<username>/role', methods=['POST'])
@admin_required
def update_user_role(username):
    new_role = (request.json or {}).get('role', '')
    if new_role not in ('admin', 'viewer'):
        return jsonify({'error': 'Role must be admin or viewer'}), 400
    if username == request.current_user:
        return jsonify({'error': 'You cannot change your own role'}), 400
    _, rows = db.fetch_all("SELECT role FROM users WHERE username = %s", (username,))
    if not rows:
        return jsonify({'error': 'User not found'}), 404
    old_role = rows[0][0]
    db.execute("UPDATE users SET role = %s WHERE username = %s", (new_role, username))
    log_change(asset_id='SYSTEM', action='UPDATE', field_changed='role',
               old_value=old_role, new_value=new_role,
               changed_by=request.current_user,
               notes=f"Role updated for user '{username}'")
    return jsonify({'success': True})


@app.route('/api/users/<username>/status', methods=['POST'])
@admin_required
def update_user_status(username):
    active = (request.json or {}).get('is_active', True)
    if username == request.current_user:
        return jsonify({'error': 'You cannot deactivate your own account'}), 400
    db.execute("UPDATE users SET is_active = %s WHERE username = %s", (active, username))
    log_change(asset_id='SYSTEM', action='UPDATE', field_changed='is_active',
               old_value=str(not active), new_value=str(active),
               changed_by=request.current_user,
               notes=f"Account {'activated' if active else 'deactivated'} for user '{username}'")
    return jsonify({'success': True})


# ── Password Reset ────────────────────────────────────────────────────────
@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    data = request.json
    identifier = data.get('identifier', '').strip()
    # Find user by username or email
    cols, rows = db.fetch_all(
        "SELECT username, email, full_name FROM users WHERE (username = %s OR email = %s) AND is_active = TRUE AND is_approved = TRUE",
        (identifier, identifier)
    )
    # Always return success to prevent enumeration
    if not rows:
        return jsonify({'success': True, 'message': 'If that account exists, a reset code has been sent.'})
    
    username, email, full_name = rows[0]
    if not email:
        return jsonify({'success': True, 'message': 'If that account exists, a reset code has been sent.'})

    code = generate_otp()
    expires = datetime.now(timezone.utc) + timedelta(minutes=15)
    db.execute(
        "UPDATE users SET reset_code = %s, reset_expires_at = %s WHERE username = %s",
        (code, expires, username)
    )
    send_email_async(send_password_reset_code, email, username, code)

    # Notify all admins
    _, admin_rows = db.fetch_all(
        "SELECT email FROM users WHERE role = 'admin' AND is_active = TRUE AND email IS NOT NULL"
    )
    for row in admin_rows:
        if row[0]:
            send_email_async(send_admin_password_reset_notification, row[0], username, email)

    return jsonify({'success': True, 'message': 'If that account exists, a reset code has been sent.'})


@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    data = request.json
    identifier = data.get('identifier', '').strip()
    code       = data.get('code', '').strip()
    new_pass   = data.get('new_password', '').strip()

    if not identifier or not code or not new_pass:
        return jsonify({'error': 'All fields are required'}), 400

    cols, rows = db.fetch_all(
        "SELECT username, reset_code, reset_expires_at FROM users WHERE (username = %s OR email = %s) AND is_active = TRUE",
        (identifier, identifier)
    )
    if not rows:
        return jsonify({'error': 'Invalid reset code'}), 401

    username, stored_code, expires_at = rows[0]
    if not stored_code or code != stored_code:
        return jsonify({'error': 'Invalid reset code'}), 401

    if expires_at and datetime.now(timezone.utc) > expires_at.replace(tzinfo=timezone.utc):
        return jsonify({'error': 'Reset code has expired'}), 401

    password_hash = bcrypt.hashpw(new_pass.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    db.execute(
        "UPDATE users SET password_hash = %s, reset_code = NULL, reset_expires_at = NULL WHERE username = %s",
        (password_hash, username)
    )
    log_change(asset_id='SYSTEM', action='UPDATE', field_changed='password',
               changed_by=username, notes='Password reset via self-service')
    return jsonify({'success': True})


# ── Delete User ────────────────────────────────────────────────────────────
@app.route('/api/users/<username>/delete', methods=['DELETE'])
@admin_required
def delete_user(username):
    if username == request.current_user:
        return jsonify({'error': 'You cannot delete your own account'}), 400
    _, rows = db.fetch_all("SELECT email FROM users WHERE username = %s", (username,))
    if not rows:
        return jsonify({'error': 'User not found'}), 404
    db.execute("DELETE FROM users WHERE username = %s", (username,))
    log_change(asset_id='SYSTEM', action='DELETE', field_changed='users',
               old_value=username, changed_by=request.current_user,
               notes=f"User '{username}' permanently deleted")
    # Notify all admins
    _, admin_rows = db.fetch_all(
        "SELECT email FROM users WHERE role = 'admin' AND is_active = TRUE AND email IS NOT NULL AND username != %s",
        (request.current_user,)
    )
    for row in admin_rows:
        if row[0]:
            send_email_async(send_account_deleted_notification, row[0], username, request.current_user)
    return jsonify({'success': True})


# ── Pending Changes (Maker-Checker) ────────────────────────────────────────
@app.route('/api/pending', methods=['GET'])
@login_required
def list_pending():
    """List all pending changes. Admins see all; viewers see their own."""
    if request.current_role == 'admin':
        cols, rows = db.fetch_all("""
            SELECT id, asset_id, action, proposed_data, field_changed,
                   old_value, new_value, submitted_by, submitted_at, status,
                   reviewed_by, reviewed_at, review_comment
            FROM pending_changes WHERE status = 'Pending' ORDER BY submitted_at DESC
        """)
    else:
        cols, rows = db.fetch_all("""
            SELECT id, asset_id, action, proposed_data, field_changed,
                   old_value, new_value, submitted_by, submitted_at, status,
                   reviewed_by, reviewed_at, review_comment
            FROM pending_changes WHERE submitted_by = %s ORDER BY submitted_at DESC
        """, (request.current_user,))
    return jsonify([dict(zip(cols, [str(v) if v is not None else '' for v in row])) for row in rows])


@app.route('/api/pending/<change_id>/approve', methods=['POST'])
@admin_required
def approve_change(change_id):
    comment = (request.json or {}).get('comment', '')
    edited_value = (request.json or {}).get('edited_value', None)
    cols, rows = db.fetch_all(
        "SELECT asset_id, action, proposed_data, field_changed, old_value, new_value, submitted_by FROM pending_changes WHERE id = %s AND status = 'Pending'",
        (change_id,)
    )
    if not rows:
        return jsonify({'error': 'Pending change not found'}), 404

    asset_id, action, proposed_data, field_changed, old_value, new_value, submitted_by = rows[0]
    proposed = proposed_data if isinstance(proposed_data, dict) else (json.loads(proposed_data) if proposed_data else {})

    # Admin can override the proposed value
    if edited_value is not None and field_changed:
        new_value = edited_value

    # Apply the change
    if action == 'INSERT':
        # Validate the proposed payload up front. The DB has CHECK and NOT
        # NULL constraints that would reject malformed data anyway, but
        # catching it here returns a clearer HTTP 400 to the admin rather
        # than a 500 from psycopg2.
        required = ['asset_id', 'system_name', 'data_category', 'data_owner', 'data_classification']
        for field in required:
            if not clean(proposed.get(field)):
                return jsonify({'error': f"Cannot approve: proposed asset is missing required field '{field}'"}), 400
        valid_classifications = {'Public', 'Internal', 'Restricted', 'Confidential'}
        if clean(proposed.get('data_classification')) not in valid_classifications:
            return jsonify({'error': f"Cannot approve: invalid data_classification"}), 400

        # Check whether the asset_id already exists. If it does and is soft-
        # deleted, we restore it with the new data rather than failing with a
        # UniqueViolation. If it exists and is still active, refuse with a
        # clear message so the admin can decide what to do.
        proposed_asset_id = clean(proposed.get('asset_id'))
        _, existing_rows = db.fetch_all(
            "SELECT is_active FROM assets WHERE asset_id = %s",
            (proposed_asset_id,)
        )
        is_restore = False
        if existing_rows:
            if existing_rows[0][0]:
                return jsonify({
                    'error': f"Cannot approve: an active asset with ID '{proposed_asset_id}' already exists. "
                             f"Reject this pending change or ask the submitter to use Edit instead."
                }), 409
            is_restore = True

        if is_restore:
            # Restore the soft-deleted row by UPDATEing all fields and re-
            # activating it. Audit history of the original create and delete
            # remain in audit_log.
            db.execute("""
                UPDATE assets SET
                    system_name = %s, data_category = %s, description = %s,
                    business_function = %s, data_owner = %s, custodian = %s,
                    data_classification = %s, data_location = %s, who_has_access = %s,
                    access_mechanism = %s, security_controls = %s,
                    encryption_in_transit = %s, encryption_at_rest = %s,
                    retention_period = %s, backup_retention = %s, disposal_method = %s,
                    primary_use = %s, policy_reference = %s, control_mapping = %s,
                    last_review_date = %s, remarks = %s,
                    data_sensitivity = %s, business_criticality = %s,
                    risk_rating = %s, key_risks = %s,
                    inherent_likelihood = %s, inherent_impact = %s,
                    residual_likelihood = %s, residual_impact = %s,
                    risk_treatment_plan = %s, review_frequency = %s,
                    target_review_date = %s, project_manager = %s,
                    network_ports_protocols = %s,
                    is_active = TRUE,
                    created_by = %s
                WHERE asset_id = %s
            """, (
                clean(proposed.get('system_name')),
                clean(proposed.get('data_category')), clean(proposed.get('description')),
                clean(proposed.get('business_function')), clean(proposed.get('data_owner')),
                clean(proposed.get('custodian')), clean(proposed.get('data_classification')),
                clean(proposed.get('data_location')), clean(proposed.get('who_has_access')),
                clean(proposed.get('access_mechanism')), clean(proposed.get('security_controls')),
                clean(proposed.get('encryption_in_transit')), clean(proposed.get('encryption_at_rest')),
                clean(proposed.get('retention_period')), clean(proposed.get('backup_retention')),
                clean(proposed.get('disposal_method')), clean(proposed.get('primary_use')),
                clean(proposed.get('policy_reference')), clean(proposed.get('control_mapping')),
                clean(proposed.get('last_review_date')), clean(proposed.get('remarks')),
                clean(proposed.get('data_sensitivity')), clean(proposed.get('business_criticality')),
                clean(proposed.get('risk_rating')), clean(proposed.get('key_risks')),
                clean_int(proposed.get('inherent_likelihood'), 1, 5),
                clean_int(proposed.get('inherent_impact'), 1, 5),
                clean_int(proposed.get('residual_likelihood'), 1, 5),
                clean_int(proposed.get('residual_impact'), 1, 5),
                clean(proposed.get('risk_treatment_plan')),
                clean(proposed.get('review_frequency')),
                clean(proposed.get('target_review_date')),
                clean(proposed.get('project_manager')),
                clean(proposed.get('network_ports_protocols')),
                request.current_user,
                proposed_asset_id
            ))
        else:
            # Standard INSERT path. Must match the column structure of
            # add_asset (admin path) exactly.
            # 35 asset fields + created_by = 36 columns.
            # clean() coerces empty strings to None so CHECK constraints on
            # enum columns (encryption_in_transit, encryption_at_rest, etc.)
            # are not violated when viewers leave optional fields blank.
            # clean_int() validates the 1-5 range for risk-scoring fields.
            db.execute("""
                INSERT INTO assets (
                    asset_id, system_name, data_category, description,
                    business_function, data_owner, custodian, data_classification, data_location,
                    who_has_access, access_mechanism, security_controls,
                    encryption_in_transit, encryption_at_rest,
                    retention_period, backup_retention, disposal_method,
                    primary_use, policy_reference, control_mapping,
                    last_review_date, remarks,
                    data_sensitivity, business_criticality, risk_rating, key_risks,
                    inherent_likelihood, inherent_impact, residual_likelihood, residual_impact,
                    risk_treatment_plan, review_frequency, target_review_date,
                    project_manager, network_ports_protocols,
                    created_by
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                clean(proposed.get('asset_id')), clean(proposed.get('system_name')),
                clean(proposed.get('data_category')), clean(proposed.get('description')),
                clean(proposed.get('business_function')), clean(proposed.get('data_owner')),
                clean(proposed.get('custodian')), clean(proposed.get('data_classification')),
                clean(proposed.get('data_location')), clean(proposed.get('who_has_access')),
                clean(proposed.get('access_mechanism')), clean(proposed.get('security_controls')),
                clean(proposed.get('encryption_in_transit')), clean(proposed.get('encryption_at_rest')),
                clean(proposed.get('retention_period')), clean(proposed.get('backup_retention')),
                clean(proposed.get('disposal_method')), clean(proposed.get('primary_use')),
                clean(proposed.get('policy_reference')), clean(proposed.get('control_mapping')),
                clean(proposed.get('last_review_date')), clean(proposed.get('remarks')),
                clean(proposed.get('data_sensitivity')), clean(proposed.get('business_criticality')),
                clean(proposed.get('risk_rating')), clean(proposed.get('key_risks')),
                clean_int(proposed.get('inherent_likelihood'), 1, 5),
                clean_int(proposed.get('inherent_impact'), 1, 5),
                clean_int(proposed.get('residual_likelihood'), 1, 5),
                clean_int(proposed.get('residual_impact'), 1, 5),
                clean(proposed.get('risk_treatment_plan')),
                clean(proposed.get('review_frequency')),
                clean(proposed.get('target_review_date')),
                clean(proposed.get('project_manager')),
                clean(proposed.get('network_ports_protocols')),
                request.current_user
            ))

    elif action == 'UPDATE' and asset_id and field_changed:
        # Allow-list check — field_changed is read from the database and
        # interpolated directly into the UPDATE statement because psycopg2
        # cannot parameterise identifiers. Without this check, any row in
        # pending_changes with a malicious field_changed value would cause
        # SQL injection at approval time.
        if field_changed not in UPDATABLE_ASSET_COLUMNS:
            return jsonify({'error': f'Invalid field: {field_changed}'}), 400
        db.execute(
            f"UPDATE assets SET {field_changed} = %s WHERE asset_id = %s",
            (new_value, asset_id)
        )

    elif action == 'DELETE':
        db.execute("UPDATE assets SET is_active = FALSE WHERE asset_id = %s", (asset_id,))

    # Mark pending change as approved
    db.execute("""
        UPDATE pending_changes SET status = 'Approved', reviewed_by = %s,
        reviewed_at = NOW(), review_comment = %s WHERE id = %s
    """, (request.current_user, comment, change_id))

    # Log to audit trail
    log_change(asset_id=asset_id, action=action,
               field_changed=field_changed, old_value=old_value, new_value=new_value,
               changed_by=request.current_user,
               notes=f'Approved pending change {change_id} submitted by {submitted_by}')

    return jsonify({'success': True})


@app.route('/api/pending/<change_id>/reject', methods=['POST'])
@admin_required
def reject_change(change_id):
    comment = (request.json or {}).get('comment', '')
    _, rows = db.fetch_all(
        "SELECT asset_id, submitted_by FROM pending_changes WHERE id = %s AND status = 'Pending'",
        (change_id,)
    )
    if not rows:
        return jsonify({'error': 'Pending change not found'}), 404

    db.execute("""
        UPDATE pending_changes SET status = 'Rejected', reviewed_by = %s,
        reviewed_at = NOW(), review_comment = %s WHERE id = %s
    """, (request.current_user, comment, change_id))

    log_change(asset_id=rows[0][0], action='REJECT',
               changed_by=request.current_user,
               notes=f'Rejected pending change {change_id}. Reason: {comment}')

    return jsonify({'success': True})




# ── Serve frontend ─────────────────────────────────────────────────────────
@app.route('/')
def index():
    with open('templates/index.html', 'r', encoding='utf-8') as f:
        return f.read()


if __name__ == '__main__':
    app.run(debug=True, port=5000)