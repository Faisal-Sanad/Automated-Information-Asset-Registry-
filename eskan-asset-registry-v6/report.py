import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import db
from datetime import datetime
import os


# ── Colour palette ─────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
GREEN       = "E2EFDA"
RED         = "FFDAD6"
YELLOW      = "FFF2CC"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"


def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)


def _set_col_widths(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_header_row(ws, row_num, headers, fill_hex, font_colour=WHITE, font_size=10):
    fill = _header_fill(fill_hex)
    font = Font(bold=True, color=font_colour, size=font_size)
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()


# ── Sheet 1: Asset Register ────────────────────────────────────────────────
def sheet_asset_register(wb):
    ws = wb.create_sheet("Asset Register")
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:U1")
    title = ws["A1"]
    title.value = "Information Asset Register — CBB OM-5.5 Compliant"
    title.font = Font(bold=True, size=13, color=WHITE)
    title.fill = _header_fill(DARK_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Asset ID",
        "Asset Name",
        "Data Category",
        "Description",
        "Business Function",
        "Data Owner",
        "Data Custodian",
        "Data Classification",
        "Data Location / System Application",
        "Access (Who)",
        "Access Method",
        "Security Controls",
        "Encryption in Transit",
        "Encryption at Rest",
        "Retention Period",
        "Backup Retention",
        "Disposal Method",
        "Permitted Use",
        "Policy Reference",
        "CBB / ISO Control Mapping",
        "Last Review Date",
        "Remarks",
        "Data Sensitivity",
        "Business Criticality",
        "Risk Rating",
        "Key Risks"
    ]
    _write_header_row(ws, 2, headers, MID_BLUE)
    ws.row_dimensions[2].height = 32

    cols, rows = db.fetch_all("""
        SELECT asset_id, system_name, data_category, description,
               business_function, data_owner, custodian, data_classification, data_location,
               who_has_access, access_mechanism, security_controls,
               encryption_in_transit, encryption_at_rest, retention_period,
               backup_retention, disposal_method, primary_use,
               policy_reference, control_mapping, last_review_date, remarks,
               data_sensitivity, business_criticality, risk_rating, key_risks
        FROM assets WHERE is_active = TRUE ORDER BY asset_id
    """)

    classification_colours = {
        "Restricted":   "FFDAD6",
        "Confidential": "FFF2CC",
        "Internal":     "E2EFDA",
        "Public":       "FFFFFF"
    }

    for r_idx, row in enumerate(rows, start=3):
        classification = str(row[6]) if row[6] else ""
        row_colour = classification_colours.get(classification, WHITE)
        row_fill = PatternFill("solid", fgColor=row_colour)

        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(value) if value else "")
            cell.fill = row_fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _border()

    widths = [10, 22, 16, 28, 20, 20, 14, 22, 22, 18, 24, 12, 10, 14, 14, 18, 24, 12, 22, 12, 20, 28, 14, 14, 12, 28]
    _set_col_widths(ws, widths)


# ── Sheet 2: Compliance Dashboard ──────────────────────────────────────────
def sheet_compliance_dashboard(wb):
    ws = wb.create_sheet("Compliance Dashboard")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Compliance Dashboard — Register Completeness"
    title.font = Font(bold=True, size=13, color=WHITE)
    title.fill = _header_fill(DARK_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Asset ID", "System Name", "Classification", "Data Category",
               "Last Review", "Missing Fields", "Compliance Status", "Review Status"]
    _write_header_row(ws, 2, headers, MID_BLUE)
    ws.row_dimensions[2].height = 28

    cols, rows = db.fetch_all("SELECT * FROM compliance_dashboard")

    for r_idx, row in enumerate(rows, start=3):
        compliance_status = str(row[6]) if row[6] else ""
        review_status     = str(row[7]) if row[7] else ""

        if compliance_status == "Compliant":
            status_fill = PatternFill("solid", fgColor=GREEN)
        else:
            status_fill = PatternFill("solid", fgColor=RED)

        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(value) if value else "")
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()

            if c_idx == 7:
                cell.fill = status_fill
                cell.font = Font(bold=True, size=9,
                                 color="375623" if compliance_status == "Compliant" else "9C0006")
            elif c_idx == 8 and review_status != "Up to Date":
                cell.fill = PatternFill("solid", fgColor=YELLOW)
            else:
                cell.fill = PatternFill("solid", fgColor=GREY if r_idx % 2 == 0 else WHITE)

    _set_col_widths(ws, [14, 34, 14, 18, 12, 16, 18, 16])

    # Summary box
    summary_row = len(rows) + 4
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=10)
    cols2, summary = db.fetch_all("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN compliance_status = 'Compliant' THEN 1 ELSE 0 END) AS compliant,
            SUM(CASE WHEN compliance_status = 'Non-Compliant' THEN 1 ELSE 0 END) AS non_compliant
        FROM compliance_dashboard
    """)
    if summary:
        total, compliant, non_compliant = summary[0]
        ws.cell(row=summary_row+1, column=1, value=f"Total Assets: {total}")
        ws.cell(row=summary_row+2, column=1, value=f"Compliant: {compliant}")
        ws.cell(row=summary_row+3, column=1, value=f"Non-Compliant: {non_compliant}")


# ── Sheet 3: Reconciliation Findings ───────────────────────────────────────
def sheet_reconciliation(wb, run_id=None):
    ws = wb.create_sheet("Reconciliation Findings")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Reconciliation Findings"
    title.font = Font(bold=True, size=13, color=WHITE)
    title.fill = _header_fill(DARK_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Run ID", "Asset ID", "Finding Type", "Detail", "CSV Value", "Register Value", "Resolved"]
    _write_header_row(ws, 2, headers, MID_BLUE)
    ws.row_dimensions[2].height = 28

    if run_id:
        cols, rows = db.fetch_all("""
            SELECT run_id, asset_id, finding_type, detail, csv_value, register_value, resolved
            FROM discrepancies WHERE run_id = %s ORDER BY finding_type
        """, (run_id,))
    else:
        cols, rows = db.fetch_all("""
            SELECT run_id, asset_id, finding_type, detail, csv_value, register_value, resolved
            FROM discrepancies ORDER BY finding_type
        """)

    type_colours = {
        "Undocumented":   "FFDAD6",
        "Misclassified":  "FFF2CC",
        "Missing Fields": "E2EFDA"
    }

    for r_idx, row in enumerate(rows, start=3):
        finding_type = str(row[2]) if row[2] else ""
        row_colour = type_colours.get(finding_type, WHITE)

        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(value) if value else "")
            cell.fill = PatternFill("solid", fgColor=row_colour)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _border()

    _set_col_widths(ws, [38, 14, 16, 40, 18, 18, 10])


# ── Sheet 4: Audit Log ─────────────────────────────────────────────────────
def sheet_audit_log(wb):
    ws = wb.create_sheet("Audit Log")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Audit Log — All Changes"
    title.font = Font(bold=True, size=13, color=WHITE)
    title.fill = _header_fill(DARK_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Timestamp", "Asset ID", "Action", "Field Changed", "Old Value", "New Value", "Changed By"]
    _write_header_row(ws, 2, headers, MID_BLUE)
    ws.row_dimensions[2].height = 28

    cols, rows = db.fetch_all("""
        SELECT changed_at, asset_id, action, field_changed, old_value, new_value, changed_by
        FROM audit_log ORDER BY changed_at DESC
    """)

    action_colours = {"INSERT": GREEN, "UPDATE": YELLOW, "DELETE": RED}

    for r_idx, row in enumerate(rows, start=3):
        action = str(row[2]) if row[2] else ""
        row_colour = action_colours.get(action, WHITE)

        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(value) if value else "")
            cell.fill = PatternFill("solid", fgColor=row_colour)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _border()

    _set_col_widths(ws, [28, 14, 10, 18, 28, 28, 16])


# ── Main ───────────────────────────────────────────────────────────────────
def generate_report(run_id=None, output_dir="."):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"Asset_Registry_Report_{timestamp}.xlsx"
    filepath  = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    print("  Building Asset Register sheet ...")
    sheet_asset_register(wb)

    print("  Building Compliance Dashboard sheet ...")
    sheet_compliance_dashboard(wb)

    print("  Building Reconciliation Findings sheet ...")
    sheet_reconciliation(wb, run_id=run_id)

    print("  Building Audit Log sheet ...")
    sheet_audit_log(wb)

    wb.save(filepath)
    print(f"\n  Report saved: {filepath}\n")
    return filepath


if __name__ == "__main__":
    import sys
    run_id = sys.argv[1] if len(sys.argv) > 1 else None
    print(f"\n{'='*60}")
    print(f"  Generating Excel Report — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")
    generate_report(run_id=run_id)